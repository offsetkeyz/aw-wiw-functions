import datetime
from hashlib import new
from typing import Pattern
from openpyxl.cell import cell
from openpyxl.descriptors.base import Integer, String
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Color, Alignment, PatternFill, GradientFill, Border, Side, NamedStyle
from dateutil.tz import *
import pytz

# Calculates normal pay perks for 12N and 12D
#   Then, calculates STAT pay for days and nights. If a shift qualifies for STAT, it is entered and set to 0.
#   Then, calculates OT for remaining shifts. There will be no duplicate shifts since STAT shifts are now 0.
# Night shifts are counted the day before OT and STAT days, thus the additional variable for nights.


schedule_wb = load_workbook('/Users/colin.mcallister/Library/CloudStorage/OneDrive-ArcticWolfNetworksInc/Documents/triage_schedule_from_wiw.xlsx')
perks_workbook = load_workbook('/Users/colin.mcallister/Documents/PERKS with Python/iSOC Perks.xlsx')
perks_ws = perks_workbook['iSOC Perks']

### MUST BE 2 DIGITS
US_stat_dates = []
# calculates stat for night before STAT holiday (add day and day after dates)
US_stat_nights = []
US_OT_days = []
US_OT_nights = []

CA_stat_dates = []
CA_stat_nights = []
CA_OT_days = []
CA_OT_nights = []

all_perks_names = []
all_perks_cells = []
all_isoc_names = []


date_range = [datetime.datetime(2022, 7, 17, tzinfo=tzutc()), datetime.datetime(2022, 7, 30, tzinfo=tzutc())]



#strips leading and trailing spaces
def strip_spaces(name_in):
    if type(name_in) is not str:
        return name_in
    if name_in != None:
        name_in_stripped = name_in.strip()
        return name_in_stripped 
    return name_in


def is_canadian(employee_in):
    if find_name_in_perks(employee_in.name) < 120:
        return True
    return False

def compare_names():
    for i in range(0,4,1):
        for cell in schedule_wb.worksheets[i]['A']:
            try:
                if len(cell.value) > 1 and '>' not in str(cell.value):
                    all_isoc_names.append(cell.value)
            except Exception as e:
                continue
    for cell in perks_workbook['iSOC Perks']['D']:
        try:
            if len(cell.value) > 1 and '>' not in str(cell.value):
                all_perks_names.append(cell.value)
                all_perks_cells.append(cell)
        except Exception as e:
            continue

    for name in all_isoc_names:
        if name not in all_perks_names:
            print(name)
                

# given the contents of the cell, returns an INT with how many hours worked
def cell_hours_worked(cell_value):
    if cell_value == None:
        return 0
    cell_value = strip_spaces(cell_value)
    if cell_value == '':
        return 0
    cell_value = str(cell_value)
    if 'V' in str(cell_value):
        return 0
    if 'S' in str(cell_value):
        return 0
    if cell_value[0].isdigit():
        if len(cell_value) > 1:
            if cell_value[1].isdigit():
                try:
                    return int(cell_value[:2])
                except:
                    print('error with cell_hours_worked')
        try:
            return int(cell_value[0])
        except:
            print('error with cell_hours_worked')
    return 0

# custom object for each employee with attributes for night hours, weekend hours, and meals          
class employee:
    name = ""
    night_hours = 0
    weekend_hours = 0
    meals = 0
    overtime = 0
    row = 0
    week1_hours = 0
    week2_hours = 0
    shifts_worked = {}
    stat_hours = 0


    def __init__(self, name, night_hours, weekend_hours, meals, overtime, row, shifts_worked, stat_hours_in):
        self.name = name
        self.night_hours = night_hours
        self.weekend_hours = weekend_hours
        self.meals = meals
        self.overtime = overtime
        self.row = row
        self.shifts_worked = shifts_worked
        self.stat_hours = stat_hours_in



#------------------ Main Functions ----------------------#

# Add numbers into correct column. Function takes name, night hours, 12H day hours
    #search for name and return row number. (cell.row)
def find_name_in_perks(name_in) -> Integer:
    for i in range(len(all_perks_names)):
        if strip_spaces(name_in) == strip_spaces(all_perks_names[i]):
            return all_perks_cells[i].row
    return -1

def calc_percs_by_section():
    all_names = {} # names : column number
    date_columns = {}

    #stores columns of dates in global dict. Columns same for all sheets
    current_ws = schedule_wb['TSE1']
    for cell in current_ws[1]:
        date_columns[str(cell.value)] = cell.column 

    for sheet in schedule_wb.worksheets:
        sheet_names = {}
        for cell in sheet['A']:
            if cell.coordinate != 'A1':                
                sheet_names[cell.value] = cell.row
        all_names[sheet.title] = sheet_names
        for row in sheet.iter_rows(2, sheet.max_row, min_col=date_columns[datetime.datetime.strftime(date_range[0], '%d %b %Y')], max_col=date_columns[datetime.datetime.strftime(date_range[1], '%d %b %Y')]):
            new_employee = employee(name="", night_hours=0, weekend_hours=0, meals=0,overtime=0, row=row[0].row, shifts_worked={}, stat_hours_in=0)
            new_employee.name = sheet.cell(row=row[0].row, column = 1).value
            for cell in row:
                cell_value = cell.value
                check_and_calc_for_12s(new_employee, cell_value)
                if cell_value == None:
                    cell_value =0
                new_employee.shifts_worked[sheet.cell(row=1, column=cell.column).value] = cell_value
            calculate_stat(new_employee)
            calculate_OT(new_employee)
            add_to_perks(new_employee)

    perks_workbook.save('/Users/colin.mcallister/Documents/PERKS with Python/iSOC Perks July17-July30.xlsx')

def calculate_OT(employee_in):
    if is_canadian(employee_in):
        if len(CA_OT_days) > 0:
            for i in CA_OT_days:
                current_value = str(employee_in.shifts_worked.get(i))
                employee_OT = employee_in.overtime
                hours_worked = cell_hours_worked(current_value)
                employee_OT += hours_worked
                employee_in.overtime = employee_OT
                employee_in.shifts_worked[i] = 0
            if len(CA_OT_nights) > 0:
                for j in CA_OT_nights:
                    current_value = str(employee_in.shifts_worked.get(j))
                    if '12N' in current_value:
                        employee_OT = employee_in.overtime
                        employee_OT += cell_hours_worked(current_value)
                        employee_in.overtime = employee_OT
                        employee_in.shifts_worked[j] = 0
    else:
        if len(US_OT_days) > 0:
            for i in US_OT_days:
                current_value = str(employee_in.shifts_worked.get(i))
                employee_OT = employee_in.overtime
                hours_worked = cell_hours_worked(current_value)
                employee_OT += hours_worked
                employee_in.overtime = employee_OT
                employee_in.shifts_worked[i] = 0
            if len(US_OT_nights) > 0:
                for j in US_OT_nights:
                    current_value = str(employee_in.shifts_worked.get(j))
                    if '12N' in current_value:
                        employee_OT = employee_in.overtime
                        employee_OT += cell_hours_worked(current_value)
                        employee_in.overtime = employee_OT
                        employee_in.shifts_worked[j] = 0
    

def calculate_stat(employee_in):
    # worked_xmas_eve = False
    if is_canadian(employee_in):
        if len(CA_stat_dates) > 0:
            for i in CA_stat_dates:
                current_value = str(employee_in.shifts_worked.get(i))
                employee_stat = employee_in.stat_hours
                hours_worked = cell_hours_worked(current_value)
                employee_stat += hours_worked
                # if i==24 and hours_worked > 0:
                #     worked_xmas_eve = True
                # elif i==25 and worked_xmas_eve:
                #     continue # go back to top if this is christmas and they already got credit for eve
                employee_in.stat_hours = employee_stat
                # stat has been calculated so it removes this value before calculating OT
                employee_in.shifts_worked[i] = 0

            if len(CA_stat_nights) > 0:
                for j in CA_stat_nights:
                    current_value = str(employee_in.shifts_worked.get(j))
                    if '12N' in current_value:
                        employee_stat = employee_in.stat_hours
                        employee_stat += cell_hours_worked(current_value)
                        employee_in.stat_hours = employee_stat
                        employee_in.shifts_worked[j] = 0
                #TODO Calc OT
    else: #if US employee
        if len(US_stat_dates) > 0:
            # worked_xmas_eve = False
            for i in US_stat_dates:

                current_value = str(employee_in.shifts_worked.get(i))
                employee_stat = employee_in.stat_hours
                hours_worked = cell_hours_worked(current_value)
                # if i==24 and hours_worked > 0:
                #     worked_xmas_eve = True
                # elif i==25 and worked_xmas_eve:
                #     continue # go back to top if this is christmas and they already got credit for eve
                employee_stat += hours_worked
                employee_in.stat_hours = employee_stat
                employee_in.shifts_worked[i] = 0

            if len(US_stat_nights) > 0:
                for j in US_stat_nights:
                    current_value = str(employee_in.shifts_worked.get(j))

                    if '12N' in current_value:
                        employee_stat = employee_in.stat_hours
                        employee_stat += cell_hours_worked(current_value)
                        employee_in.stat_hours = employee_stat
                        employee_in.shifts_worked[i] = 0



 #used by above function. this does the counting    
def check_and_calc_for_12s(new_employee, cell_value):
    if str(cell_value).strip() == "12D":
        new_employee.weekend_hours = new_employee.weekend_hours + 12
        new_employee.meals = new_employee.meals + 1
    if str(cell_value).strip() == "12N":
        new_employee.night_hours = new_employee.night_hours + 12
        new_employee.meals = new_employee.meals + 1
    elif str(cell_value).__contains__('N') and len(str(cell_value)) <=3:
        if len(str(cell_value)) == 3:
            new_employee.night_hours = new_employee.night_hours + int(cell_value[:2])
        else:
            new_employee.night_hours = new_employee.night_hours + int(cell_value[0])

def add_to_perks(new_employee): 
    # add data to perks WS
    if new_employee.name is not None:
        perks_row = find_name_in_perks(new_employee.name)
        if perks_row < 0:
            return
        perks_ws.cell(row=perks_row, column=5).value = new_employee.night_hours
        perks_ws.cell(row=perks_row, column=7).value = new_employee.weekend_hours
        perks_ws.cell(row=perks_row, column=9).value = new_employee.meals
        perks_ws.cell(row=perks_row, column=11).value = new_employee.overtime
        perks_ws.cell(row=perks_row, column=12).value = new_employee.stat_hours


def main():
    compare_names()
    input("DID YOU CHANGE THE DATE RANGE?")
    calc_percs_by_section()

if __name__ == '__main__':
    main()
    