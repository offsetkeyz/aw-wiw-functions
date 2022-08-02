import csv
import datetime
from lib2to3.pgen2 import token
import bs_methods
from openpyxl import load_workbook

workbook_location = '/Users/colin.mcallister/Downloads/Time_Off_Request-2022-08-01.150217.xlsx'
wb=load_workbook(workbook_location)
current_ws = wb['Time Off Request']

token = bs_methods.authenticate_WiW_API()
all_users = bs_methods.get_all_wiw_users(token)

bs_methods.create_time_off_request(token, all_users[bs_methods.get_user_id_from_email(token, 'stephanie.legue@arcticwolf.com')], "2022-07-01T10:10:10Z", "2022-08-03T10:10:10Z")
stuffs = []
for cell in current_ws['C']:
    if cell.row < 3:
        continue
    if current_ws.cell(row=cell.row, column=9).value == '--':
        continue
    if current_ws.cell(row=cell.row, column=5).value == 'Approved':
        name = str(cell.value).split(', ')
        email = name[1].replace(' ','').replace('-','') + '.' + name[0].replace(' ','').replace('-','') + '@arcticwolf.com'
        user_id =bs_methods.get_user_id_from_email(token, email)
        # user = all_users[user_id]
        try:
            user = all_users[user_id]
        except Exception as e:
            stuffs.append(str(e))
            continue
        start_date = datetime.datetime.strptime(current_ws.cell(row=cell.row, column=8).value, '%M-%d-%Y')
        end_date = datetime.datetime.strptime(current_ws.cell(row=cell.row, column=9).value, '%M-%d-%Y')
        if start_date == end_date:
            end_date = end_date + datetime.timedelta(days=1)
        formatted_start = start_date.strftime("%Y-%m-%dT%H:%M:%SZ")
        formatted_end = end_date.strftime("%Y-%m-%dT%H:%M:%SZ")
        _ = bs_methods.create_time_off_request(token, user, formatted_start, formatted_end)
    
with open('time_off.csv', 'w') as csvfile:
    csvwriter = csv.writer(csvfile)
    csvwriter.writerows(stuffs)

# for user in all_users: #Key = user_id and value = user object
#     all_requests = bs_methods.get_time_off_requests_for_user(token, all_users[user].wiw_employee_id)
#     for request in all_requests[user]:
#         bs_methods.create_time_off_request(token, all_users[user], str(request.start_time.strftime("%Y-%m-%dT%H:%M:%SZ")), request.end_time.strftime("%Y-%m-%dT%H:%M:%SZ"))