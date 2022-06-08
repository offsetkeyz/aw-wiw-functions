from ctypes import alignment
from datetime import *
from distutils.command import build
import json
from lib2to3.pgen2 import token
from logging.config import DEFAULT_LOGGING_CONFIG_PORT
import sched
from secrets import token_bytes
from sqlite3 import enable_shared_cache
from tkinter import CENTER, HORIZONTAL, VERTICAL
from xmlrpc.client import DateTime
from dateutil.tz import *
import requests
import bs_methods
import shift_classes
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment
import pytz

workbook_location = '/Users/colin.mcallister/Library/CloudStorage/OneDrive-ArcticWolfNetworksInc/Documents/triage_schedule_from_wiw.xlsx'
workbook = load_workbook(workbook_location)
date_columns = {}
all_names = {}
all_positions = {
    'Business Analyst 1':10762839,
    'Business Systems Manager':10762840,
    'Co-op/ Intern':10762841,
    'Concierge Sec Eng 3':10762842,
    'Concierge Sec Engineer 2':10762843,
    'Concierge Sec Engnr 2': 10762843,
    'Concierge Security Eng 2':10762843,
    'Concierge Sec Engnr 3':10762842,
    'Concierge Security Eng 3':10762842,
    'Dir, Security Services':10762845,
    'IT Applications Admin':10762846,
    'Mgr Concierge Services':10762847,
    'Mgr Technical Operations':10762848,
    'Mgr, Concierge Security':10762847,
    'Mgr, Security Operations':10762849,
    'Network Ops Supp Analyst':10762850,
    'Network Sec Ops Analyst 2':10762850,
    'Network Sec Ops Anlst 1':10762850,
    'S2 Technical Trainer 4':10762851,
    'Shift Lead Sec Ops':10762852,
    'Shift Lead Security Ops':10762852,
    'Sr Dir, Sec Ops':10762854,
    'Sr Mgr, Security Ops':10762855,
    'SVP, Security Services':10762856,
    'Tech Lead Security Svcs':10762857,
    'Triage Sec Eng 1':10762858,
    'Triage Security Engr 1':10762858,
    'Triage Security Eng 1':10762858,
    'Triage Security Analyst':10762858,
    'Triage Security Eng 2':10762859,
    'Triage Security Engr 2':10762859,
    'Triage Sec Eng 3':10762860,
    'Triage Security Eng 3':10762860,
    'Triage Security Engr 3':10762860,
    'Triage Security Engr 4':10762861,
    'VP, Business Applications':10762862,
    'CAN': 10762911,
    'USA': 10762910,
    'DEU':10762912,
    'DNK':10762912,
    'GBR':10762912
}

def list_all_wiw_users(token):
    all_users = []
    url_headers = bs_methods.get_url_and_headers('users', token)
    response = requests.request("GET", url_headers[0], headers=url_headers[1])
    for i in response.json()['users']:
        all_users.append(shift_classes.user(i['first_name'], i['last_name'], i['email'], i['employee_code'], i['positions'], i['role'], i['locations'], i['is_hidden'], i['is_active']))
    return all_users

#takes in location (schedule) ID and returns which worksheet
def check_location_id(location_id) -> str:
    schedules = {
        #5129876 : "Default"
        5132409 : "TSE1",
        5132410 : "TSE2",
        5134192 : "TSE3",
        5132412 : "TechOps",
        # 5189759 : "Colin Test",
        5227330 : "EMEATier1",
        # 5129876 : "Pink",
        5233779 : "EMEATier3"
    }
    try:
        return schedules[int(location_id)]
    except:
        return 0

def build_date_row():
    for sheet in workbook.sheetnames:
        current_ws = workbook[sheet]
        start_date = datetime(2022,2,28)
        for col in current_ws.iter_cols(min_row=1, max_row=1, min_col=3, max_col=360):
            for cell in col:
                cell.value = datetime.strftime(start_date, '%d %b %Y')
                date_columns[datetime.strftime(start_date, '%d %b %Y')] = cell.column #stores columns of dates in global dict. Columns same for all sheets
                cell.fill = PatternFill("solid", fgColor="FFEFC2")
                if datetime.strftime(datetime.strptime(cell.value, '%d %b %Y'), '%a').startswith('S'): #weekend
                    cell.fill = PatternFill("solid", fgColor="AFD5FF")
            start_date = start_date + timedelta(days=1)
        workbook.save(str(workbook_location))

def isoc_team_structure_update(token):
    s2_employees = get_iSOC_employees()
    structure_users = get_isoc_structure_users()

    # highlight if fields are missing
    # if team member is TSE1 or TSA, highlight if den is missing.


def get_isoc_structure_users():
    iSOC_Team_Structure = load_workbook('/Users/colin.mcallister/Library/CloudStorage/OneDrive-ArcticWolfNetworksInc/Documents/iSOC Team Structure - Colin.xlsx', data_only=True)
    S2_Roster = iSOC_Team_Structure['Raw User Data']
    all_s2_employees = {}
    
    for row in range(2, S2_Roster.max_row):
        current_email = str(S2_Roster.cell(row=row, column=4).value).strip().lower()
        job_role = S2_Roster.cell(row=row, column=3).value
        active_status = S2_Roster.cell(row=row, column=6).value
        zendesk_id = S2_Roster.cell(row=row, column=1).value
        full_name = ''
        enable = True
        manager = ''
        all_s2_employees[current_email] = {'id' : zendesk_id, 'position' : job_role, 'status' :active_status, 'email' : current_email}

        #TODO Get Zendesk ID spreadsheet
    return all_s2_employees

def update_users(token):
    s2_employees = get_s2_employees()
    wiw_users = list_all_wiw_users(token)
    for user in wiw_users:
        try:
            user_details = s2_employees[str(user.email).strip().lower()]
        except:
            continue
        update_wiw_user(token,user_details)

def get_s2_employees():
    weekly_headcount = load_workbook('/Users/colin.mcallister/Downloads/Weekly Headcount Report - S2.xlsx', data_only=True)
    S2_Roster = weekly_headcount['Current Employee List']
    all_s2_employees = {}
    for row in range(2, S2_Roster.max_row):
        current_email = str(S2_Roster.cell(row=row, column=19).value).strip().lower()
        position = S2_Roster.cell(row=row, column=5).value
        active_status = S2_Roster.cell(row=row, column=22).value
        region = S2_Roster.cell(row=row, column=10).value
        employee_id = S2_Roster.cell(row=row, column=1).value
        manager = S2_Roster.cell(row=row, column=12).value
        all_s2_employees[current_email] = {'id' : employee_id, 'position' : position, 'status' :active_status, 'region' : region, 'email' : current_email, 'manager': manager}
    return all_s2_employees

def get_iSOC_employees():
    weekly_headcount = load_workbook('/Users/colin.mcallister/Downloads/Weekly Headcount Report - S2.xlsx', data_only=True)
    S2_Roster = weekly_headcount['S2 Roster']
    all_headcount_positions = {'Triage Sec Analyst':'TSA', 'Triage Security Eng 1':'TSE1', 'Triage Security Eng 2':'TSE2', 'Triage Security Eng 3':'TSE3', 'Triage Security Engr 1':'TSE1', 'Triage Security Engr 2':'TSE2', 'Triage Security Engr 3': 'TSE3', 'Triage Sec Analyst':'TSA', 'Triage Sec Eng 1':'TSE1', 'Triage Sec Eng 2':'TSE2', 'Triage Sec Eng 3':'TSE3', 'Triage Sec Eng 4':'TSE4', 'Triage Sec Engineer 1':'TSE1', 'Triage Security Analyst':'TSA'}
    position_acronyms = ['TSA','TSE1','TSE2','TSE3','TSE4','Intern']
    # if INTERN, don't check. Not in weekly headcount
    # highlight if position is different, or if they aren't in weekly headcount --DON'T DELETE--
    all_s2_employees = {}
    for row in range(2, S2_Roster.max_row):
        current_email = str(S2_Roster.cell(row=row, column=19).value).strip().lower()
        position = S2_Roster.cell(row=row, column=5).value
        active_status = S2_Roster.cell(row=row, column=22).value
        region = S2_Roster.cell(row=row, column=10).value
        employee_id = S2_Roster.cell(row=row, column=1).value
        manager = S2_Roster.cell(row=row, column=12).value
        all_s2_employees[current_email] = {'id' : employee_id, 'position' : position, 'status' :active_status, 'region' : region, 'email' : current_email, 'manager': manager}
    return all_s2_employees

def update_wiw_user(token,user_details): #10656558 is "unknown"    
    url_headers = bs_methods.get_url_and_headers('users/' + str(bs_methods.get_user_id_from_email(token, user_details['email'])),token)
    user_positions = []
    if user_details['position'] not in all_positions:
        user_positions.append(10656558)
        print(user_details['position'] + ", " + user_details['email'])
        all_positions[user_details['position']] = 10656558
    else: 
        user_positions.append(all_positions[user_details['position']])

    payload = json.dumps({
        "employee_code": user_details['id'],
        "positions": [
            user_positions[0],
            all_positions[user_details['region']]
        ],
    }) 
    success=False
    i = 1
    while success == False and i < 10:
        try:
            response = requests.request("PUT", url_headers[0], headers=url_headers[1], data=payload)
            success = True  
        except:
            i+=1


def get_date_rows():
    current_ws = workbook['TSE1']
    for cell in current_ws[1]:
        date_columns[str(cell.value)] = cell.column #stores columns of dates in global dict. Columns same for all sheets

# returns dict with Worksheet as key and an array of names as value
def get_all_names():
    for sheet in workbook.worksheets:
        sheet_names = {} # names : column number
        for cell in sheet['A']:
            sheet_names[cell.value] = cell.row
        all_names[sheet.title] = sheet_names

#takes in shift and returns First and Last name STRING
def get_name_from_shift(token, shift_in) -> str:
    current_user = bs_methods.get_user_from_id(token, shift_in.user_id)
    return str(current_user.first_name) + ' ' + str(current_user.last_name) 

def check_sheet_for_name(name_in, schedule_name):
    if name_in not in all_names[schedule_name]:
        #add to worksheet
        ws = workbook[schedule_name]
        current_row = ws.max_row+1
        if current_row == 2:
            current_row = 3
        ws.cell(row=current_row, column=1).value = name_in
        #add to dict
        all_names[schedule_name][name_in] = current_row

def populate_user_in_excel_sheet(user_shifts, schedule_name, user):
        if user.full_name == 'Silvia Lira-Perez':
            print("asdf")
        team_number = False
        for shift in user_shifts:
           try:
                ws = workbook[check_location_id(shift.location_id)]  
                check_sheet_for_name(user.full_name, check_location_id(shift.location_id))     
                current_cell = ws.cell(row=all_names[check_location_id(shift.location_id)][user.full_name], column=date_columns[datetime.strftime(shift.start_time, '%d %b %Y')])
           except KeyError as e:
                continue

           current_cell.value = shift.length
           if shift.location_id not in [5227330,5233779] and 0 <= int(datetime.strftime(shift.start_time, '%-H')) <= 6:
               current_cell.value = str(int(current_cell.value)) + 'N'
           elif shift.location_id not in [5227330,5233779]:
                current_cell.value = str(int(current_cell.value)) + 'D'
            #checks if the shift is within 20 days of current day to assign team number
           if team_number == False and (int(datetime.strftime(datetime.now(), '%-j'))-10) < int(datetime.strftime(shift.start_time, '%-j')) < (int(datetime.strftime(datetime.now(), '%-j'))+10):
                team_cell = ws.cell(row=all_names[schedule_name][user.full_name], column=2)
                team_cell.value = bs_methods.get_team_number(shift.site_id)
                team_number = True
           if shift.published == True:
                current_cell.fill = PatternFill("solid", fgColor=shift.color)
           else:
                current_cell.fill = PatternFill("solid", fgColor='ffffff')
                current_cell.border = Border(left=Side(border_style='thick', color=shift.color),right=Side(border_style='thick', color=shift.color),top=Side(border_style='thick', color=shift.color),bottom=Side(border_style='thick', color=shift.color))
        #    current_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),top=Side(style='thick'),bottom=Side(style='thick'))
           current_cell.comment = Comment(shift.notes, "iSOC Scheduling")

def populate_users_time_off(user_requests, schedule_name, user):
        for request in user_requests:
           ws = workbook[schedule_name]
           date_check = request.start_time
           while request.start_time <= date_check < request.end_time:
                try:
                    current_cell = ws.cell(row=all_names[schedule_name][user.full_name], column=date_columns[datetime.strftime(date_check, '%d %b %Y')])
                except KeyError as e:
                    date_check = date_check + timedelta(days=1)
                    continue                
                current_cell.value = 'V - Time Off'
                current_cell.fill = PatternFill("solid", fgColor='ff8789')
                current_cell.comment = Comment(request.type_label, "iSOC Scheduling")
                date_check = date_check + timedelta(days=1)

def get_time_off_requests(token):
    url_headers = bs_methods.get_url_and_headers('requests', token)
    response = requests.request("GET", url_headers[0], headers=url_headers[1]).json()
    requests_json = response['requests']
    return bs_methods.store_time_off(requests_json)
    

def create_today_hyperlinks():
    for i in workbook.sheetnames:
        ws = workbook[i]
        column_letter=str(ws.cell(row=1, column=date_columns[datetime.strftime(datetime.now(), '%d %b %Y')]).column_letter)
        display_text = '#' + i + '!' + column_letter + str(ws.min_row) + ':' + column_letter + str(ws.max_row)
        ws['A2'].hyperlink = display_text
        ws['A2'].value = ">> TODAY <<"
        ws['A2'].style = "Hyperlink"
        ws['A2'].font = Font(bold=True, size=20)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A2'].border = Border(left=Side(style='double'), right=Side(style='double'),top=Side(style='double'),bottom=Side(style='double'))

def completely_clear_sheets():
    for i in workbook.sheetnames:
        ws = workbook[i]
        ws.delete_rows(2, ws.max_row)

def clear_future_columns(start_date:datetime):
    get_date_rows()
    for i in workbook.sheetnames:
        ws = workbook[i]
        ws.delete_cols(date_columns[datetime.strftime(start_date, '%d %b %Y')], ws.max_column)
    build_date_row()
    get_date_rows()

def main():
    token = bs_methods.authenticate_WiW_API()
    # isoc_team_structure_update(token)
    clear_future_columns(datetime.now()-timedelta(days=21))
    build_date_row()
    get_date_rows()
    get_all_names()    
    update_users(token)
    all_to_requests = get_time_off_requests(token)
    all_shifts_json = bs_methods.get_all_shifts_json(token)
    all_shifts = bs_methods.store_shifts_by_user_id(all_shifts_json) #returns dict with user_id as key
    for user_id in all_shifts:
        user = bs_methods.get_user_from_id(token, user_id)
        user_shifts = all_shifts[user_id]
        schedule_name = check_location_id(user_shifts[0].location_id)
        if schedule_name == 0:
            continue #shift not on a tracked schedule
        check_sheet_for_name(user.full_name, schedule_name)
        populate_user_in_excel_sheet(user_shifts, schedule_name, user)
        try:
            populate_users_time_off(all_to_requests[user_id],schedule_name, user)
        except Exception as e:
            _ = e
            continue
    # update_users(token)
    create_today_hyperlinks()
    workbook.save(str(workbook_location))


if __name__ == "__main__":
    main()

